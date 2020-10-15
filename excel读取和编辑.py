# 引入第三方库openpyxl,把excel文件加到项目目录中
import pprint
import openpyxl
import requests
# 获取工作表中的数据
def workdata(excelName,workName):
    wk = openpyxl.open(filename=excelName)  # 获取excel文件
    sheet = wk[workName]  # 获取excel文件中的工作表
    sheet.cell(row=2, column=1).value  # 获取该工作表中第2行第1列的值
    # print(sheet.cell(row=2, column=1).value)
    maxrow = sheet.max_row  # 工作表最大行数
    print("最大行数：", maxrow)
    maxcolumn=sheet.max_column  # 工作表最大列数
    # 读取工作表中每行的数据
    list1 = []
    for num in range(2, maxrow + 1):
        case_id = sheet.cell(row=num, column=1).value  # 获取工作表中每行的case_id
        url = sheet.cell(row=num, column=5).value  # 获取工作表中每行的url
        data = sheet.cell(row=num, column=6).value  # 获取工作表中每行的传递数据
        expected = sheet.cell(row=num, column=7).value  # 获取工作表中每行的预期值
        case = dict(case_id=case_id, url=url, data=data, expected=expected)  # 把每行的数据存储到字典中
        # print("打印每行的数据：", case)
        list1.append(case)  # 把所有的数据存储到list中
    # print("当前工作表的列表:", list1)
    return list1

# print("获取工作表中所有的数据列表：",workdatas)
# pprint.pprint(workdatas)
'''
打印结果
{'case_id': 1,
  'data': '{"mobile_phone":"13552440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}',
  'expected': '{"code": 0, "msg": "OK"}',
  'url': 'http://api.lemonban.com/futureloan/member/register'},
'''
# 修改工作表中的数据
def updatework(excelname,workname,row,result):
    wkedit = openpyxl.open(excelname)
    editsheet = wkedit[workname]
    editsheet.cell(row=row, column=8).value = result # 给每行的第8列赋值
    wkedit.save(excelname)  # 保存execl表格
# post请求
def postdata(workurl,workdata):
    header={"X-Lemonban-Media-Type":"lemonban.v2",
    "Content-Type":"application/json"}
    # 调用post函数
    response=requests.post(url=workurl,json=workdata,headers=header) #由于这里是json格式的，所以用的是json参数
    res=response.json()  #获取调用post函数返回的值
    # pprint.pprint(res)
    return res;
# 执行用例
def Bj(excelname,workname,id=''):
    workdatas = workdata(excelname,workname)  #调用工作表取得数据，返回list列表
    print("长度：",len(workdatas))
    for workone in workdatas:
        workurl=workone['url'] #获取工作表中的url
        workcase_id=workone['case_id']
        if id=='':
            json=eval(workone['data'])  #处理字符串表达式，返回表达式的值，由于取到的data是字符串，所以用eval函数来转化为字典，它也可以转化为元组、列表
        else:
            json = eval(workone['data'])
            json['member_id']=id
        workexpected=eval(workone['expected'])['msg']
        pprint.pprint(json)
        resdata=postdata(workurl, json)  #调用返回的数据
        print("结果：",workexpected)
        if workexpected==resdata['msg']:  #判断工作表中的预期结果和返回的实际结果
            updatework(excelname, workname, workcase_id+1, '通过')  #调用修改工作表函数
            print(f"用例{workcase_id}通过")
            print('*'*35)
        else:
            updatework(excelname, workname, workcase_id + 1, '不通过')
            print(f"用例{workcase_id}不通过")
            print('*'*35)
        return resdata
Bj("test_case_api.xlsx",'register')
loginresponse=Bj("test_case_api.xlsx",'login')
Bj("test_case_api.xlsx",'recharge',loginresponse['data']['id'])


