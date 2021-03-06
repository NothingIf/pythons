﻿'''
自动化测试的流程
1、编写测试用例  ===ok
2、读取excel测试用例数据===ok
3、通过requests发送请求===ok
4、比对实际结果与预期结果是否一致 ====回写执行的结果：结果一致则通过，否则不通过
'''


import pprint
import openpyxl
import requests


# 读取excel数据
def ReadData(filename,sheetname):
    # 获取工作簿

    wk= openpyxl.open(filename=filename)
    # 获取当前工作表
    sheet1=wk[sheetname]
    # print(sheet1.cell(row=1,column=1).value)
    # 获取最大行数
    # print('获取最大行数：',sheet1.max_row)
    # 获取最大列数
    # print('获取最大列数:',sheet1.max_column)

    # # 如何取得excel所有的数据\
    datalist=[]
    for x in range(2,sheet1.max_row+1):
        # 把用例数据存储到字典
        # print(sheet1.cell(row=x, column=1).value)
        case=dict(case_id=sheet1.cell(x,1).value,
            url=sheet1.cell(row=x,column=5).value,
             data=sheet1.cell(row=x,column=6).value,
             expected=sheet1.cell(row=x,column=7).value)
        # print(case)
        datalist.append(case)
        #是否需要返回值
    return datalist

# 发送请求
def api_request(url,json):
    url=url
    json=json
    headers={"X-Lemonban-Media-Type":"lemonban.v2",
    "Content-Type":"application/json"}
    response= requests.post(url=url, json=json, headers=headers)
    return response.json()

# 回写测试执行的结果
def write_result(filename,sheetname,row,column,result):
    wk=openpyxl.open(filename)
    sheet1=wk[sheetname]
    sheet1.cell(row,column).value=result
    wk.save(filename)

# 执行测试用例
def exec_case(filename,sheetname):
    # 获取excel所有用例的数据
    cases=ReadData(filename,sheetname)
    '''
    每一条用例请求数据：
    {'case_id': 1,
    'url': 'http://api.lemonban.com/futureloan/member/register',
    'data': '{"mobile_phone":"13552440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}',
     'expected': '{"code": 0, "msg": "OK"}'
     }
    '''
    for case in cases:
        # print('每一条用例请求数据：')
        # print(case)
        case_id=case["case_id"]
        url=case['url']
        ## 去引号把data数据转换为字典
        # print(case.get("data"),type(case.get("data")))
        data=eval(case.get("data"))
        # 发送请求获取响应的数据
        response=api_request(url=url,json=data)
        # print('用例的响应json数据：')
        # print(response)
        #比对实际结果和预期结果
        # 获取预期结果
        excepted_msg=eval(case["expected"])["msg"]
        print(f'用例预期结果:',excepted_msg)
        #获取期望结果
        real_msg=response["msg"]
        print(f'用例实际结果:', real_msg)
        if real_msg==excepted_msg:
            print(f"用例{case_id}测试执行通过")
            write_result(filename=filename,sheetname=sheetname,
                         row=case_id+1,column=8,result='通过')
        else:
            print(f"用例{case_id}测试执行不通过")
            write_result(filename=filename, sheetname=sheetname,
                         row=case_id+1, column=8, result='不通过')
        print('*'*30)

# 执行注册用例
exec_case('test_case_api.xlsx',"register")
# 执行登录用例
exec_case('test_case_api.xlsx',"login")

# eval函数的应用
# '''
#     eval 处理字符串表达式，返回表达式的值，它也可以把字符串转化为字典、元组、列表
#     eval('3*2')
# '''
# a=eval('(1,2,4,5)')
# print(type(a),a)


